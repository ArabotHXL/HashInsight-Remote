import json
import logging
import os
import secrets
import io
import csv
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import FastAPI, HTTPException, UploadFile, File, Request
from . import __version__
from fastapi.responses import HTMLResponse, JSONResponse, PlainTextResponse, Response, FileResponse
from fastapi.staticfiles import StaticFiles

from .config import (
    AppConfig,
    MinerConfig,
    load_config,
    save_config,
    get_config_path,
    get_last_warnings,
    DEFAULT_LOCAL_PORT,
    MIN_LATEST_INTERVAL_SEC,
    MIN_RAW_INTERVAL_SEC,
)

# ---------------------------
# Local API authentication (UI -> Local API)
# ---------------------------
LOCAL_SECRET_HEADER = "X-Local-API-Secret"
LOCAL_SECRET_ENV = "PICKAXE_LOCAL_API_SECRET"
LOCAL_SECRET_MIN_LEN = 16

def _get_local_secret(cfg_obj) -> str:
    try:
        sec = getattr(cfg_obj, "local_api_secret", "") or ""
    except Exception:
        sec = ""
    sec = (sec or "").strip()
    if not sec:
        sec = (os.getenv(LOCAL_SECRET_ENV, "") or "").strip()
    return sec

def _is_loopback(req: Request) -> bool:
    try:
        host = req.client.host if req.client else ""
    except Exception:
        host = ""
    return host in ("127.0.0.1", "::1", "localhost")

def _require_local_secret(req: Request):
    cfg_dict = load_config()
    cfg_obj = AppConfig(**cfg_dict) if isinstance(cfg_dict, dict) else cfg_dict
    secret = _get_local_secret(cfg_obj)
    if not secret:
        raise HTTPException(status_code=400, detail="Local API secret is not configured")
    provided = (req.headers.get(LOCAL_SECRET_HEADER, "") or "").strip()
    if not provided:
        auth = (req.headers.get("Authorization", "") or "").strip()
        if auth.lower().startswith("bearer "):
            provided = auth[7:].strip()
    if provided != secret:
        raise HTTPException(status_code=401, detail="Unauthorized")

def _safe_config_payload(cfg_obj) -> dict:
    d = cfg_obj.__dict__.copy() if hasattr(cfg_obj, "__dict__") else dict(cfg_obj)
    d.pop("local_api_secret", None)
    d["local_api_secret_configured"] = bool(_get_local_secret(cfg_obj))
    return d
from .logging_setup import setup_logging
from .runtime import CollectorRunner
from .vendor_edge_collector.cgminer_client import CGMinerClient, CGMinerError
from .vendor_edge_collector.ip_scanner import IPRangeParser


# Optional dependency used for multipart/form-data uploads in FastAPI.
# If not installed, we keep the app runnable and disable the file-upload API.
try:
    import multipart  # type: ignore

    HAS_MULTIPART = True
except Exception:
    HAS_MULTIPART = False


def _normalize_miner_rows(rows: List[List[str]], defaults: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Normalize CSV/XLSX rows into miner dicts.

    Supported inputs:
      - 1 column: ip or ip:port
      - 2 columns: miner_id, ip
      - 3+ columns: miner_id, ip, port, type
      - With header row containing 'ip'/'miner_id' etc.

    Returns list of dicts: {miner_id, ip, port, miner_type}
    """
    def clean(x: Any) -> str:
        return str(x).strip() if x is not None else ""

    # Drop entirely empty rows.
    rows2: List[List[str]] = []
    for r in rows or []:
        rr = [clean(x) for x in (r or [])]
        if any(v for v in rr):
            rows2.append(rr)
    if not rows2:
        return []

    # Detect header.
    header = [c.lower() for c in rows2[0]]
    has_header = any(h in ("ip", "miner_id", "miner", "port", "type", "miner_type") for h in header)

    col_map = {}
    if has_header:
        for i, h in enumerate(header):
            h2 = h.strip().lower()
            if h2 in ("miner_id", "id", "miner"):
                col_map["miner_id"] = i
            elif h2 in ("ip", "host"):
                col_map["ip"] = i
            elif h2 == "port":
                col_map["port"] = i
            elif h2 in ("type", "miner_type"):
                col_map["miner_type"] = i
        data_rows = rows2[1:]
    else:
        data_rows = rows2

    out: List[Dict[str, Any]] = []
    auto = 0
    default_port = int(defaults.get("port", 4028))
    default_type = str(defaults.get("miner_type", defaults.get("type", "antminer")) or "antminer")
    id_prefix = str(defaults.get("id_prefix", "AUTO_") or "AUTO_")

    for r in data_rows:
        # If we have a header map, use it.
        if col_map:
            ip = clean(r[col_map.get("ip", 0)]) if col_map.get("ip") is not None else ""
            miner_id = clean(r[col_map.get("miner_id", -1)]) if col_map.get("miner_id") is not None and col_map.get("miner_id", -1) >= 0 else ""
            port_raw = clean(r[col_map.get("port", -1)]) if col_map.get("port") is not None and col_map.get("port", -1) >= 0 else ""
            typ = clean(r[col_map.get("miner_type", -1)]) if col_map.get("miner_type") is not None and col_map.get("miner_type", -1) >= 0 else ""
        else:
            # Positional fallback.
            rr = [clean(x) for x in r]
            # If row looks like: ip,port,type (from Excel) (no miner_id)
            if len(rr) >= 2 and rr[0] and rr[1].isdigit() and (len(rr) < 3 or rr[2] in ("antminer", "whatsminer", "avalon", "innosilicon", "goldshell", "other", "")):
                miner_id = ""
                ip = rr[0]
                port_raw = rr[1]
                typ = rr[2] if len(rr) >= 3 else ""
            else:
                miner_id = rr[0] if len(rr) >= 1 else ""
                ip = rr[1] if len(rr) >= 2 else rr[0]
                port_raw = rr[2] if len(rr) >= 3 else ""
                typ = rr[3] if len(rr) >= 4 else ""

        # Support ip:port in ip field
        if ip and ":" in ip:
            a, b = ip.split(":", 1)
            ip = a.strip()
            if (not port_raw) and b.strip().isdigit():
                port_raw = b.strip()

        if not ip:
            continue

        try:
            port = int(port_raw) if port_raw else default_port
        except Exception:
            port = default_port

        miner_type = typ or default_type
        if not miner_id:
            auto += 1
            miner_id = f"{id_prefix}{str(auto).zfill(6)}"

        out.append({
            "miner_id": miner_id,
            "ip": ip,
            "port": port,
            "miner_type": miner_type,
        })

    return out

logger = logging.getLogger("PickaxeLocalAPI")


def _data_dir() -> Path:
    return get_config_path(None).parent


def _log_dir() -> Path:
    return _data_dir() / "logs"


def tail_file(path: Path, lines: int = 200) -> str:
    if not path.exists():
        return ""
    try:
        with path.open("rb") as f:
            f.seek(0, os.SEEK_END)
            size = f.tell()
            # Read last ~64KB for tail (enough for 200 lines typically)
            block = 65536
            seek = max(0, size - block)
            f.seek(seek)
            data = f.read().decode("utf-8", errors="ignore")
        out = data.splitlines()[-lines:]
        return "\n".join(out)
    except Exception:
        return ""


def create_app() -> FastAPI:
    dd = _data_dir()
    log_file = setup_logging(_log_dir())

    runner = CollectorRunner(dd)

    app = FastAPI(title="HashInsight Pickaxe Collector", version=__version__)

    web_dir = Path(__file__).parent / "web"
    static_dir = web_dir / "static"
    static_dir.mkdir(parents=True, exist_ok=True)
    # Serve only static assets (JS/CSS) from /static
    app.mount("/static", StaticFiles(directory=str(static_dir)), name="static")

    @app.get("/", response_class=HTMLResponse)
    def index() -> str:
        return (web_dir / "index.html").read_text(encoding="utf-8")

    # Backward-compatible asset route. Older UIs referenced /app.js directly.
    # The canonical path is now /static/app.js.
    @app.get("/app.js")
    def legacy_app_js() -> Response:
        return FileResponse(static_dir / "app.js", media_type="application/javascript")

    # Backward-compatible alias. Some earlier UIs referenced /app_js.
    @app.get("/app_js")
    def legacy_app_js_alias() -> Response:
        return legacy_app_js()


    # --- Local API Secret bootstrap ---
    @app.get("/api/local-secret")
    def api_local_secret_status():
        cfg_dict = load_config()
        cfg_obj = AppConfig(**cfg_dict) if isinstance(cfg_dict, dict) else cfg_dict
        return {"configured": bool(_get_local_secret(cfg_obj))}

    @app.post("/api/local-secret")
    def api_local_secret_set(payload: Dict[str, Any], req: Request):
        # Restrict bootstrap to loopback.
        if not _is_loopback(req):
            raise HTTPException(status_code=403, detail="Forbidden")

        cfg_dict = load_config()
        cfg_obj = AppConfig(**cfg_dict) if isinstance(cfg_dict, dict) else cfg_dict
        current = _get_local_secret(cfg_obj)

        # If already configured, require existing secret to rotate.
        if current:
            provided = (req.headers.get(LOCAL_SECRET_HEADER, "") or "").strip()
            if provided != current:
                raise HTTPException(status_code=401, detail="Unauthorized")

        if bool(payload.get("generate")):
            new_secret = secrets.token_urlsafe(32)
        else:
            new_secret = str(payload.get("secret") or "").strip()
            if len(new_secret) < LOCAL_SECRET_MIN_LEN:
                raise HTTPException(status_code=400, detail=f"Secret too short (min {LOCAL_SECRET_MIN_LEN})")

        cfg_obj.local_api_secret = new_secret
        save_config(cfg_obj)
        return {"configured": True, "secret": new_secret}

    @app.get("/api/version")
    def version() -> Dict[str, Any]:
        return {
            "app": "HashInsight Pickaxe Collector",
            "version": __version__,
            "config_path": str(get_config_path(None)),
            "log_file": str(log_file),
        }

    @app.get("/api/capabilities")
    def capabilities() -> Dict[str, Any]:
        return {
            "multipart_upload": HAS_MULTIPART,
            "excel_import": True,
        }

    if HAS_MULTIPART:

        @app.post("/api/miners/import_file")
        async def api_import_miners_file(
            file: UploadFile = File(...),
            default_port: int = 4028,
            default_type: str = "antminer",
            id_prefix: str = "AUTO_",
        ) -> Dict[str, Any]:
            """Parse a CSV/XLSX file into miner objects for Bulk Add.

            Expected columns (any order, header optional): miner_id, ip, port, type
            If only a single column is present, it's treated as ip or ip:port.
            """
            name = (file.filename or "").lower()
            content = await file.read()
            if not content:
                raise HTTPException(status_code=400, detail="Empty file")

            defaults = {"port": default_port, "type": default_type, "id_prefix": id_prefix}

            rows: List[List[str]] = []
            try:
                if name.endswith(".csv") or name.endswith(".txt"):
                    text = content.decode("utf-8", errors="ignore")
                    reader = csv.reader(io.StringIO(text))
                    for r in reader:
                        rows.append([c for c in r])
                elif name.endswith(".xlsx") or name.endswith(".xlsm") or name.endswith(".xltx"):
                    try:
                        from openpyxl import load_workbook
                    except Exception as e:
                        raise HTTPException(status_code=500, detail=f"openpyxl not available: {e}")

                    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
                    ws = wb.active
                    for r in ws.iter_rows(values_only=True):
                        rows.append(["" if v is None else str(v) for v in r])
                else:
                    raise HTTPException(status_code=400, detail="Unsupported file type. Use .csv or .xlsx")
            except HTTPException:
                raise
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Failed to parse file: {e}")

            miners = _normalize_miner_rows(rows, defaults)
            return {"success": True, "count": len(miners), "miners": miners}

    else:

        @app.post("/api/miners/import_file")
        async def api_import_miners_file_unavailable() -> Dict[str, Any]:
            raise HTTPException(
                status_code=501,
                detail='File upload requires "python-multipart". Install it with: pip install python-multipart',
            )

    @app.get("/api/config")
    def api_get_config() -> Dict[str, Any]:
        cfg = load_config(None)
        return {
            "warnings": get_last_warnings(),
            "config": {
                "site_id": cfg.site_id,
                "site_name": cfg.site_name,
                "cloud_api_base": cfg.cloud_api_base,
                "collector_token": cfg.collector_token,
                "zone_id": getattr(cfg, "zone_id", ""),
                "device_id": getattr(cfg, "device_id", ""),
                "telemetry_api_mode": getattr(cfg, "telemetry_api_mode", "legacy"),
                "command_api_mode": getattr(cfg, "command_api_mode", "auto"),
                "ack_include_snapshot": getattr(cfg, "ack_include_snapshot", True),
                # v0.2+: two-loop polling strategy
                "latest_interval_sec": cfg.latest_interval_sec,
                "raw_interval_sec": cfg.raw_interval_sec,
                # backward compatible alias (UI may still show poll_interval_sec)
                "poll_interval_sec": cfg.poll_interval_sec,
                "timeout_sec": cfg.timeout_sec,
                "max_retries": cfg.max_retries,
                "max_workers": cfg.max_workers,
                "batch_size": cfg.batch_size,
                "upload_connect_timeout_sec": cfg.upload_connect_timeout_sec,
                "upload_read_timeout_sec": cfg.upload_read_timeout_sec,
                "upload_workers": cfg.upload_workers,
                "latest_max_miners": cfg.latest_max_miners,
                "shard_total": cfg.shard_total,
                "shard_index": cfg.shard_index,
                "miner_timeout_fast_sec": cfg.miner_timeout_fast_sec,
                "miner_timeout_slow_sec": cfg.miner_timeout_slow_sec,
                "offline_backoff_base_sec": cfg.offline_backoff_base_sec,
                "offline_backoff_max_sec": cfg.offline_backoff_max_sec,
                "enable_commands": cfg.enable_commands,
                "command_poll_interval_sec": cfg.command_poll_interval_sec,
                "upload_ip_to_cloud": cfg.upload_ip_to_cloud,
                "encrypt_miners_config": cfg.encrypt_miners_config,
                "local_key_env": cfg.local_key_env,
                # Binding store (CSV -> SQLite) - local-only inventory
                "binding_enable": getattr(cfg, "binding_enable", True),
                "binding_csv_path": getattr(cfg, "binding_csv_path", "./miners.csv"),
                "binding_db_path": getattr(cfg, "binding_db_path", ""),
                "binding_encrypt_credentials": getattr(cfg, "binding_encrypt_credentials", True),
                # Offline spool retention
                "offline_spool_max_age_hours": getattr(cfg, "offline_spool_max_age_hours", 24),
                "offline_spool_max_total_bytes": getattr(cfg, "offline_spool_max_total_bytes", 10 * 1024 * 1024 * 1024),
                # Burst sampling
                "enable_burst_sampling": getattr(cfg, "enable_burst_sampling", False),
                "burst_interval_sec": getattr(cfg, "burst_interval_sec", 10),
                "burst_duration_sec": getattr(cfg, "burst_duration_sec", 300),
                "burst_hashrate_drop_pct": getattr(cfg, "burst_hashrate_drop_pct", 15),
                "burst_temp_threshold_c": getattr(cfg, "burst_temp_threshold_c", 85),
                # Privacy toggles
                "mask_ip_in_logs": getattr(cfg, "mask_ip_in_logs", True),
                "enable_whatsminer_http": getattr(cfg, "enable_whatsminer_http", True),
                "inventory_sources": list(getattr(cfg, "inventory_sources", ["miners","binding","ip_ranges"])),
                "miners": [m.__dict__ for m in cfg.miners],
                "ip_ranges": cfg.ip_ranges,
            },
            "limits": {
                "min_latest_interval_sec": MIN_LATEST_INTERVAL_SEC,
                "min_raw_interval_sec": MIN_RAW_INTERVAL_SEC,
            },
        }

    @app.post("/api/config")
    def api_save_config(payload: Dict[str, Any], req: Request) -> Dict[str, Any]:
        raw = payload.get("config")
        if not isinstance(raw, dict):
            raise HTTPException(status_code=400, detail="Missing config")

        _require_local_secret(req)

        miners_raw = raw.get("miners", []) or []
        miners: List[MinerConfig] = []
        for m in miners_raw:
            if not m:
                continue
            try:
                miners.append(MinerConfig(
                    miner_id=str(m.get("miner_id") or m.get("id") or m.get("ip")),
                    ip=str(m.get("ip")),
                    port=int(m.get("port", 4028)),
                    miner_type=str(m.get("miner_type") or m.get("type") or "antminer"),
                ))
            except Exception:
                continue

        # Handle older UI payloads that only provide poll_interval_sec.
        poll = int(raw.get("poll_interval_sec", 60))
        latest = int(raw.get("latest_interval_sec", poll))
        raw_int = int(raw.get("raw_interval_sec", max(latest, 60)))
        latest = max(MIN_LATEST_INTERVAL_SEC, latest)
        raw_int = max(MIN_RAW_INTERVAL_SEC, raw_int)

        shard_total = max(1, int(raw.get("shard_total", 1)))
        shard_index = int(raw.get("shard_index", 0))
        if shard_index < 0 or shard_index >= shard_total:
            shard_index = 0

        current_cfg_dict = load_config()
        current_cfg = AppConfig(**current_cfg_dict) if isinstance(current_cfg_dict, dict) else current_cfg_dict
        current_secret = getattr(current_cfg, "local_api_secret", "")

        cfg = AppConfig(
            site_id=str(raw.get("site_id", "site_001")),
            site_name=str(raw.get("site_name", "")),
            cloud_api_base=str(raw.get("cloud_api_base", "")),
            collector_token=str(raw.get("collector_token", "")),
            zone_id=str(raw.get("zone_id", getattr(current_cfg, "zone_id", ""))),
            device_id=str(raw.get("device_id") or getattr(current_cfg, "device_id", "")),
            telemetry_api_mode=str(raw.get("telemetry_api_mode") or getattr(current_cfg, "telemetry_api_mode", "legacy")),
            command_api_mode=str(raw.get("command_api_mode") or getattr(current_cfg, "command_api_mode", "auto")),
            ack_include_snapshot=raw.get("ack_include_snapshot", getattr(current_cfg, "ack_include_snapshot", True)) in (True, "true", "True", 1, "1"),
            local_api_secret=str(current_secret or ""),
            latest_interval_sec=latest,
            raw_interval_sec=raw_int,
            poll_interval_sec=latest,
            timeout_sec=float(raw.get("timeout_sec", 5)),
            max_retries=int(raw.get("max_retries", 5)),
            max_workers=int(raw.get("max_workers", 50)),
            batch_size=int(raw.get("batch_size", 1000)),
            upload_connect_timeout_sec=float(raw.get("upload_connect_timeout_sec", 2)),
            upload_read_timeout_sec=float(raw.get("upload_read_timeout_sec", 30)),
            upload_workers=int(raw.get("upload_workers", 4)),
            latest_max_miners=int(raw.get("latest_max_miners", 500)),
            shard_total=shard_total,
            shard_index=shard_index,
            miner_timeout_fast_sec=float(raw.get("miner_timeout_fast_sec", 1.5)),
            miner_timeout_slow_sec=float(raw.get("miner_timeout_slow_sec", 5.0)),
            offline_backoff_base_sec=int(raw.get("offline_backoff_base_sec", 30)),
            offline_backoff_max_sec=int(raw.get("offline_backoff_max_sec", 300)),
            enable_commands=bool(raw.get("enable_commands", False)),
            command_poll_interval_sec=int(raw.get("command_poll_interval_sec", 5)),
            upload_ip_to_cloud=raw.get("upload_ip_to_cloud", False) in (True, "true", "True", 1, "1"),
            encrypt_miners_config=raw.get("encrypt_miners_config", False) in (True, "true", "True", 1, "1"),
            local_key_env=str(raw.get("local_key_env", "PICKAXE_LOCAL_KEY")),
            binding_enable=raw.get("binding_enable", getattr(current_cfg, "binding_enable", True)) in (True, "true", "True", 1, "1"),
            binding_csv_path=str(raw.get("binding_csv_path", getattr(current_cfg, "binding_csv_path", "./miners.csv"))),
            binding_db_path=str(raw.get("binding_db_path", getattr(current_cfg, "binding_db_path", ""))),
            binding_encrypt_credentials=raw.get("binding_encrypt_credentials", getattr(current_cfg, "binding_encrypt_credentials", True)) in (True, "true", "True", 1, "1"),
            offline_spool_max_age_hours=int(raw.get("offline_spool_max_age_hours", getattr(current_cfg, "offline_spool_max_age_hours", 24))),
            offline_spool_max_total_bytes=int(raw.get("offline_spool_max_total_bytes", getattr(current_cfg, "offline_spool_max_total_bytes", 10 * 1024 * 1024 * 1024))),
            enable_burst_sampling=raw.get("enable_burst_sampling", getattr(current_cfg, "enable_burst_sampling", False)) in (True, "true", "True", 1, "1"),
            burst_interval_sec=int(raw.get("burst_interval_sec", getattr(current_cfg, "burst_interval_sec", 10))),
            burst_duration_sec=int(raw.get("burst_duration_sec", getattr(current_cfg, "burst_duration_sec", 300))),
            burst_hashrate_drop_pct=int(raw.get("burst_hashrate_drop_pct", getattr(current_cfg, "burst_hashrate_drop_pct", 15))),
            burst_temp_threshold_c=int(raw.get("burst_temp_threshold_c", getattr(current_cfg, "burst_temp_threshold_c", 85))),
            mask_ip_in_logs=raw.get("mask_ip_in_logs", getattr(current_cfg, "mask_ip_in_logs", True)) in (True, "true", "True", 1, "1"),
            enable_whatsminer_http=raw.get("enable_whatsminer_http", getattr(current_cfg, "enable_whatsminer_http", True)) in (True, "true", "True", 1, "1"),
            inventory_sources=list(raw.get("inventory_sources", getattr(current_cfg, "inventory_sources", ["miners","binding","ip_ranges"]))),
            miners=miners,
            ip_ranges=list(raw.get("ip_ranges", [])),
        )

        p = save_config(cfg, None)
        return {"success": True, "config_path": str(p)}

    @app.post("/api/miners/test")
    def api_test_miners(payload: Dict[str, Any]) -> Dict[str, Any]:
        miners = payload.get("miners")
        if not isinstance(miners, list) or not miners:
            raise HTTPException(status_code=400, detail="miners must be a non-empty list")

        # IMPORTANT: Testing 500 miners sequentially is unusably slow.
        # We run probes concurrently with a capped worker pool.
        timeout = float(payload.get("timeout_sec", 2.5))
        try:
            concurrency = int(payload.get("concurrency", 50))
        except Exception:
            concurrency = 50
        concurrency = max(1, min(concurrency, 200))

        # Optional sampling to protect the farm network during UI testing.
        # If absent, we test all miners.
        max_to_test = payload.get("max_to_test")
        if isinstance(max_to_test, int) and max_to_test > 0:
            miners = miners[:max_to_test]

        def _probe(m: Dict[str, Any]) -> Dict[str, Any]:
            ip = str(m.get("ip", "")).strip()
            if not ip:
                return {"ip": "", "port": 0, "alive": False, "error": "missing ip"}
            port = int(m.get("port", 4028))
            try:
                client = CGMinerClient(ip, port=port, timeout=timeout, max_retries=1)
                summary = client.get_summary()

                # Attempt to extract hashrate quickly
                ghs = None
                try:
                    s0 = (summary.get("SUMMARY") or [{}])[0]
                    if s0.get("GHS av") is not None:
                        ghs = float(s0.get("GHS av"))
                    elif s0.get("MHS av") is not None:
                        ghs = float(s0.get("MHS av")) / 1000.0
                except Exception:
                    ghs = None

                ths = (ghs / 1000.0) if (ghs is not None) else None
                return {
                    "ip": ip,
                    "port": port,
                    "alive": True,
                    "latency_ms": client.last_latency_ms,
                    # Backward compatible naming for different UIs
                    "hashrate_ghs": ghs,
                    "hashrate_ths": ths,
                }
            except CGMinerError as e:
                return {
                    "ip": ip,
                    "port": port,
                    "alive": False,
                    "error": str(e),
                }
            except Exception as e:
                return {
                    "ip": ip,
                    "port": port,
                    "alive": False,
                    "error": f"{type(e).__name__}: {e}",
                }

        results: List[Dict[str, Any]] = []
        with ThreadPoolExecutor(max_workers=concurrency) as ex:
            futures = [ex.submit(_probe, m) for m in miners]
            for fut in as_completed(futures):
                try:
                    r = fut.result()
                    if r.get("ip"):
                        results.append(r)
                except Exception as e:
                    results.append({"ip": "", "port": 0, "alive": False, "error": f"probe_failed: {e}"})

        return {"results": results, "count": len(results), "concurrency": concurrency}

    @app.post("/api/iprange/expand")
    def api_expand_range(payload: Dict[str, Any]) -> Dict[str, Any]:
        s = str(payload.get("range", "")).strip()
        if not s:
            raise HTTPException(status_code=400, detail="range is required")

        parser = IPRangeParser()
        start, end, total = parser.parse(s)
        # Hard limit for MVP safety
        max_ips = int(payload.get("max_ips", 512))
        ips = []
        for i, ip in enumerate(parser.enumerate_ips(start, end)):
            if i >= max_ips:
                break
            ips.append(ip)

        return {"start": start, "end": end, "total": total, "returned": len(ips), "ips": ips}

    @app.post("/api/collector/start")
    def api_start(req: Request) -> Dict[str, Any]:
        _require_local_secret(req)
        cfg = load_config(None)
        if not cfg.cloud_api_base or not cfg.collector_token:
            raise HTTPException(status_code=400, detail="cloud_api_base and collector_token are required")
        # Allow starting from ip_ranges-only deployments (no explicit miners list).
        if (not cfg.miners) and (not (cfg.ip_ranges or [])):
            raise HTTPException(status_code=400, detail="Please add at least one miner (or configure ip_ranges)")

        return runner.start(cfg)

    @app.post("/api/collector/stop")
    def api_stop(req: Request) -> Dict[str, Any]:
        _require_local_secret(req)
        return runner.stop()

    @app.get("/api/status")
    def api_status() -> Dict[str, Any]:
        return runner.status()

    @app.get("/api/logs", response_class=PlainTextResponse)
    def api_logs(lines: int = 200) -> str:
        return tail_file(log_file, lines=lines)

    return app
